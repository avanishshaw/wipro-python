import openpyxl

def get_excel_data(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    data = []

    for row in range(2, sheet.max_row + 1):
        username = sheet.cell(row=row, column=1).value
        password = sheet.cell(row=row, column=2).value
        first_name = sheet.cell(row=row, column=3).value
        last_name = sheet.cell(row=row, column=4).value
        postal_code = sheet.cell(row=row, column=5).value
        if username is None or password is None:
            continue

        data.append([username, password, first_name, last_name, postal_code])

    return data


def get_test_data(file_path, sheet_name):
    return get_excel_data(file_path, sheet_name)

#this is proper framework acceptable by company
